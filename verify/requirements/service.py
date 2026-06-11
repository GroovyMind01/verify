"""Requirement service — import, query, version management."""

import csv
from typing import Protocol

from sqlalchemy import select, text
from sqlalchemy.orm import selectinload

from verify.definitions.models import RequirementTestMapping
from verify.requirements.models import Requirement
from verify.shared.database import make_service_session
from verify.shared.exceptions import NotFoundError, RequirementImportError


class RequirementService(Protocol):
    """Interface for requirement operations."""

    def import_from_csv(
        self, file_path: str, update: bool = False
    ) -> list[Requirement]:
        """Import requirements from a CSV file. Returns created Requirement entities."""
        ...

    def import_from_excel(
        self, file_path: str, sheet: str | None = None, update: bool = False
    ) -> list[Requirement]:
        """Import requirements from an Excel file."""
        ...

    def list_all(self, domain: str | None = None) -> list[Requirement]:
        """List all requirements, optionally filtered by domain."""
        ...

    def get_by_key(self, key: str) -> Requirement:
        """Retrieve a requirement by its unique key."""
        ...

    def get_by_id(self, requirement_id: str) -> Requirement:
        """Retrieve a requirement by its UUID."""
        ...

    def create_version(self, key: str, updates: dict) -> Requirement:
        """Create a new version of an existing requirement."""
        ...

    def get_coverage(self, requirement_id: str) -> dict:
        """Return coverage info: which test definitions cover this requirement."""
        ...

    def search(self, query: str) -> list[Requirement]:
        """Full-text search across key, title, description."""
        ...

    def set_parent(self, child_key: str, parent_key: str) -> Requirement:
        """Set the decomposition parent of a requirement (not version chain)."""
        ...

    def diff_versions(self, key: str, v1: int, v2: int) -> dict:
        """Compare two versions of a requirement. Returns changed fields."""
        ...

    def export_markdown(self, output_path: str) -> None:
        """Export all requirements as a Markdown document."""
        ...

    def archive(self, key: str) -> Requirement:
        """Soft-delete (archive) a requirement."""
        ...

    def rebuild_fts(self) -> None:
        """Rebuild the FTS5 index."""
        ...


class RequirementServiceImpl:
    """Concrete implementation of RequirementService."""

    def __init__(self, session_factory) -> None:
        self._session_factory = lambda: make_service_session(session_factory)

    def import_from_csv(self, file_path: str, update: bool = False) -> list[Requirement]:
        with self._session_factory() as session, open(file_path, newline="") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None:
                raise RequirementImportError(f"No headers found in {file_path}")

            created = []
            for row_num, row in enumerate(reader, start=2):
                key_val = (row.get("key") or "").strip()
                if not key_val:
                    continue

                stmt = (
                    select(Requirement)
                    .where(Requirement.key == key_val)
                    .order_by(Requirement.version.desc())
                    .limit(1)
                )
                existing = session.execute(stmt).scalars().first()
                if existing is not None:
                    if update:
                        existing.title = (row.get("title") or "").strip() or existing.title
                        if row.get("description"):
                            existing.description = (row.get("description") or "").strip()
                        if row.get("domain"):
                            existing.domain = (row.get("domain") or "general").strip()
                        created.append(existing)
                    continue

                req = Requirement(
                    key=key_val,
                    title=(row.get("title") or "").strip() or "Untitled",
                    description=(row.get("description") or "").strip() or None,
                    domain=(row.get("domain") or "general").strip() or "general",
                    source_file=file_path,
                    source_line=row_num,
                )
                session.add(req)
                created.append(req)

            session.commit()
            for r in created:
                session.expunge(r)
            self.rebuild_fts()
            return created

    def import_from_excel(
        self, file_path: str, sheet: str | None = None, update: bool = False
    ) -> list[Requirement]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise RequirementImportError(f"No data in {file_path}")

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        with self._session_factory() as session:
            created = []
            for row_num, row in enumerate(rows[1:], start=2):
                data = dict(zip(headers, row))
                key_val = str(data.get("key", "")).strip() if data.get("key") else ""
                if not key_val:
                    continue

                stmt = (
                    select(Requirement)
                    .where(Requirement.key == key_val)
                    .order_by(Requirement.version.desc())
                    .limit(1)
                )
                existing = session.execute(stmt).scalars().first()
                if existing is not None:
                    if update:
                        existing.title = (
                            str(data.get("title", "")).strip() or existing.title
                        )
                        if data.get("description"):
                            existing.description = (
                                str(data.get("description", "")).strip()
                            )
                        if data.get("domain"):
                            existing.domain = (
                                str(data.get("domain", "general")).strip() or "general"
                            )
                        created.append(existing)
                    continue

                req = Requirement(
                    key=key_val,
                    title=str(data.get("title", "Untitled")).strip() or "Untitled",
                    description=str(data.get("description", "")).strip() or None,
                    domain=str(data.get("domain", "general")).strip() or "general",
                    source_file=file_path,
                    source_line=row_num,
                )
                session.add(req)
                created.append(req)

            session.commit()
            for r in created:
                session.expunge(r)
            self.rebuild_fts()
            return created

    def list_all(self, domain: str | None = None) -> list[Requirement]:
        with self._session_factory() as session:
            stmt = select(Requirement).where(Requirement.archived.is_(False))
            if domain:
                stmt = stmt.where(Requirement.domain == domain)
            results = list(session.execute(stmt.order_by(Requirement.key)).scalars())
            for r in results:
                session.expunge(r)
            return results

    def get_by_key(self, key: str) -> Requirement:
        with self._session_factory() as session:
            stmt = (
                select(Requirement)
                .where(Requirement.key == key)
                .order_by(Requirement.version.desc())
                .limit(1)
            )
            result = session.execute(stmt).scalars().first()
            if result is None:
                raise NotFoundError(f"Requirement with key '{key}' not found")
            session.expunge(result)
            return result

    def get_by_id(self, requirement_id: str) -> Requirement:
        with self._session_factory() as session:
            stmt = select(Requirement).where(Requirement.id == requirement_id)
            result = session.execute(stmt).scalar_one_or_none()
            if result is None:
                raise NotFoundError(f"Requirement with id '{requirement_id}' not found")
            session.expunge(result)
            return result

    def create_version(self, key: str, updates: dict) -> Requirement:
        with self._session_factory() as session:
            stmt = (
                select(Requirement)
                .where(Requirement.key == key)
                .order_by(Requirement.version.desc())
                .limit(1)
            )
            latest = session.execute(stmt).scalars().first()
            if latest is None:
                raise NotFoundError(f"Requirement with key '{key}' not found")

            allowed = {"title", "description", "domain", "status", "attributes"}
            new_data: dict = {
                "key": key,
                "title": latest.title,
                "description": latest.description,
                "domain": latest.domain,
                "status": latest.status,
                "attributes": latest.attributes,
                "version": latest.version + 1,
                "parent_id": latest.id,
            }
            for k, v in updates.items():
                if k in allowed:
                    new_data[k] = v

            new_req = Requirement(**new_data)
            session.add(new_req)
            session.commit()
            session.expunge(new_req)
            return new_req

    def get_coverage(self, requirement_id: str) -> dict:
        with self._session_factory() as session:
            req = session.execute(
                select(Requirement).where(Requirement.id == requirement_id)
            ).scalar_one_or_none()
            if req is None:
                raise NotFoundError(f"Requirement with id '{requirement_id}' not found")

            mappings = session.execute(
                select(RequirementTestMapping)
                .where(RequirementTestMapping.requirement_id == requirement_id)
                .options(selectinload(RequirementTestMapping.test_definition))
            ).scalars().all()

            tests = []
            for m in mappings:
                td = m.test_definition
                tests.append({
                    "test_id": td.id,
                    "test_name": td.name,
                    "coverage_claim": m.coverage_claim,
                })

            return {
                "requirement_id": req.id,
                "requirement_key": req.key,
                "requirement_title": req.title,
                "covered_by": len(tests),
                "tests": tests,
            }

    def search(self, query: str) -> list[Requirement]:
        with self._session_factory() as session:
            # Escape FTS5 special characters by quoting each term
            import re
            safe_terms = []
            for term in re.split(r"\s+", query.strip()):
                if term:
                    safe_terms.append(f'"{term}"')
            search_term = " AND ".join(safe_terms) if safe_terms else query
            raw_sql = text(
                "SELECT r.id FROM requirements r "
                "JOIN requirements_fts fts ON fts.requirement_id = r.id "
                "WHERE requirements_fts MATCH :q AND r.archived = 0"
            )
            result = session.execute(raw_sql, {"q": search_term})
            ids = [row[0] for row in result]
            if not ids:
                return []
            stmt = select(Requirement).where(Requirement.id.in_(ids))
            results = list(session.execute(stmt).scalars())
            for r in results:
                session.expunge(r)
            return results

    def set_parent(self, child_key: str, parent_key: str) -> Requirement:
        with self._session_factory() as session:
            child_stmt = (
                select(Requirement)
                .where(Requirement.key == child_key)
                .order_by(Requirement.version.desc())
                .limit(1)
            )
            child = session.execute(child_stmt).scalars().first()
            if child is None:
                raise NotFoundError(f"Requirement with key '{child_key}' not found")

            parent_stmt = (
                select(Requirement)
                .where(Requirement.key == parent_key)
                .order_by(Requirement.version.desc())
                .limit(1)
            )
            parent = session.execute(parent_stmt).scalars().first()
            if parent is None:
                raise NotFoundError(f"Requirement with key '{parent_key}' not found")

            child.decomposition_parent_id = parent.id
            session.commit()
            session.expunge(child)
            return child

    def diff_versions(self, key: str, v1: int, v2: int) -> dict:
        with self._session_factory() as session:
            ver1 = session.execute(
                select(Requirement).where(
                    Requirement.key == key, Requirement.version == v1
                )
            ).scalar_one_or_none()
            if ver1 is None:
                raise NotFoundError(f"Requirement '{key}' version {v1} not found")

            ver2 = session.execute(
                select(Requirement).where(
                    Requirement.key == key, Requirement.version == v2
                )
            ).scalar_one_or_none()
            if ver2 is None:
                raise NotFoundError(f"Requirement '{key}' version {v2} not found")

            fields = ["title", "description", "domain", "status", "attributes"]
            changes = {}
            for field in fields:
                old_val = getattr(ver1, field)
                new_val = getattr(ver2, field)
                if old_val != new_val:
                    changes[field] = (old_val, new_val)

            return {
                "key": key,
                "v1": v1,
                "v2": v2,
                "changes": changes,
            }

    def export_markdown(self, output_path: str) -> None:
        requirements = self.list_all()
        with open(output_path, "w") as f:
            f.write("# Requirements\n\n")
            for req in requirements:
                f.write(f"## {req.key}: {req.title} (v{req.version})\n\n")
                f.write(f"**Domain:** {req.domain}  \n")
                f.write(f"**Status:** {req.status}  \n")
                f.write(f"**Type:** {req.requirement_type}\n\n")
                if req.description:
                    f.write(f"{req.description}\n\n")
                f.write("---\n\n")

    def archive(self, key: str) -> Requirement:
        with self._session_factory() as session:
            stmt = (
                select(Requirement)
                .where(Requirement.key == key)
                .order_by(Requirement.version.desc())
                .limit(1)
            )
            req = session.execute(stmt).scalars().first()
            if req is None:
                raise NotFoundError(f"Requirement with key '{key}' not found")
            req.archived = True
            session.commit()
            session.expunge(req)
            return req

    def rebuild_fts(self) -> None:
        with self._session_factory() as session:
            session.execute(text("DELETE FROM requirements_fts"))
            session.execute(
                text(
                    "INSERT INTO requirements_fts(requirement_id, key, title, description) "
                    "SELECT id, key, title, COALESCE(description, '') "
                    "FROM requirements WHERE archived = 0"
                )
            )
            session.commit()
